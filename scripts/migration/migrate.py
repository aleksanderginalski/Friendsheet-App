"""
One-time migration script: Excel → Firestore.

Reads Migracja.xlsx and writes meetings, persons to Firestore under
users/{uid}/meetings and users/{uid}/persons.
Activity categories are read-only — they must already exist in Firestore.

Usage:
    python migrate.py --uid <uid> --xlsx <path> --key <service_account.json>
"""

import argparse
import datetime
import sys

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

WEIGHT_MAP = {4: 5}
BATCH_SIZE = 500


# ---------------------------------------------------------------------------
# Firebase helpers
# ---------------------------------------------------------------------------

def init_firebase(key_path: str):
    cred = credentials.Certificate(key_path)
    firebase_admin.initialize_app(cred)
    return firestore.client()


# ---------------------------------------------------------------------------
# Firestore loaders
# ---------------------------------------------------------------------------

def load_categories(db, uid: str) -> tuple[dict, dict]:
    """Return (categories_by_name, parent_by_id)."""
    docs = db.collection('users').document(uid).collection('activity_categories').stream()
    categories_by_name: dict[str, str] = {}
    parent_by_id: dict[str, str | None] = {}
    for doc in docs:
        data = doc.to_dict()
        name = data.get('name', '').strip()
        if name:
            categories_by_name[name] = doc.id
        parent_by_id[doc.id] = data.get('parentCategoryId')
    return categories_by_name, parent_by_id


def load_persons(db, uid: str) -> dict[str, str]:
    """Return persons_by_fullname: 'firstName lastName' → doc_id."""
    docs = db.collection('users').document(uid).collection('persons').stream()
    result: dict[str, str] = {}
    for doc in docs:
        data = doc.to_dict()
        first = data.get('firstName', '').strip()
        last = data.get('lastName', '').strip()
        full = f'{first} {last}'.strip() if last else first
        if full:
            result[full] = doc.id
    return result


def load_existing_meetings(db, uid: str) -> set[tuple]:
    """Return set of (date, name) for idempotency check."""
    docs = db.collection('users').document(uid).collection('meetings').stream()
    result: set[tuple] = set()
    for doc in docs:
        data = doc.to_dict()
        raw_date = data.get('date')
        name = data.get('name', '')
        if raw_date and name:
            if hasattr(raw_date, 'date'):
                result.add((raw_date.date(), name))
            else:
                # Firestore Timestamp — convert via datetime
                result.add((raw_date.ToDatetime().date(), name))
    return result


# ---------------------------------------------------------------------------
# Ancestor resolution
# ---------------------------------------------------------------------------

def get_ancestor_ids(category_id: str, parent_by_id: dict) -> list[str]:
    """Return list of ancestor doc_ids from direct parent up to root."""
    ancestors: list[str] = []
    current = parent_by_id.get(category_id)
    while current is not None:
        ancestors.append(current)
        current = parent_by_id.get(current)
    return ancestors


def resolve_category_ids(
    activity_names: list[str],
    categories_by_name: dict[str, str],
    parent_by_id: dict,
) -> list[str]:
    """
    For each activity name resolve to category_id, then collect the category
    plus all its ancestors. Returns deduplicated list, leaf-first.
    """
    seen: set[str] = set()
    result: list[str] = []

    for name in activity_names:
        cat_id = categories_by_name[name]  # KeyError propagates intentionally
        for cid in [cat_id] + get_ancestor_ids(cat_id, parent_by_id):
            if cid not in seen:
                seen.add(cid)
                result.append(cid)

    return result


# ---------------------------------------------------------------------------
# Person helpers
# ---------------------------------------------------------------------------

def split_name(full_name: str) -> tuple[str, str]:
    """'Anna Maria Kowalska' → ('Anna', 'Maria Kowalska')."""
    parts = full_name.strip().split(' ')
    first = parts[0]
    last = ' '.join(parts[1:]) if len(parts) > 1 else ''
    return first, last


def get_or_create_person(
    full_name: str,
    uid: str,
    persons_by_fullname: dict[str, str],
    new_persons: list[tuple],
) -> str:
    """Return existing doc_id or stage a new person document."""
    if full_name in persons_by_fullname:
        return persons_by_fullname[full_name]

    first, last = split_name(full_name)
    doc_id = db_ref_id()
    now = datetime.datetime.utcnow()
    doc = {
        'userId': uid,
        'firstName': first,
        'createdAt': now,
    }
    if last:
        doc['lastName'] = last

    persons_by_fullname[full_name] = doc_id
    new_persons.append((doc_id, doc))
    return doc_id


def db_ref_id() -> str:
    """Generate a Firestore-compatible random ID (20 chars, alphanumeric)."""
    import random
    import string
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=20))


# ---------------------------------------------------------------------------
# Pre-flight check
# ---------------------------------------------------------------------------

def preflight_check(
    ws,
    headers: list,
    categories_by_name: dict[str, str],
) -> None:
    """
    Scan all activity cells and verify every name exists in Firestore.
    Aborts with sys.exit(1) if any are missing.
    """
    missing: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[2]
        if not raw:
            continue
        for name in [n.strip() for n in str(raw).split(';') if n.strip()]:
            if name not in categories_by_name:
                missing.add(name)

    if missing:
        print('\nPre-flight FAILED — activities not found in Firestore:')
        for m in sorted(missing):
            print(f'  - {m}')
        print('\nAbort. No data was written.')
        sys.exit(1)

    print(f'Pre-flight OK — all activity names verified ({len(categories_by_name)} categories loaded).')


# ---------------------------------------------------------------------------
# Batch commit helper
# ---------------------------------------------------------------------------

def commit_in_batches(db, operations: list[tuple]) -> None:
    """
    operations: list of (collection_ref, doc_id, data_dict)
    Commits in chunks of BATCH_SIZE.
    """
    for start in range(0, len(operations), BATCH_SIZE):
        chunk = operations[start:start + BATCH_SIZE]
        batch = db.batch()
        for col_ref, doc_id, data in chunk:
            batch.set(col_ref.document(doc_id), data)
        batch.commit()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Migrate Excel meetings to Firestore.')
    parser.add_argument('--uid', required=True, help='Firebase user UID')
    parser.add_argument('--xlsx', required=True, help='Path to Migracja.xlsx')
    parser.add_argument('--key', required=True, help='Path to serviceAccountKey.json')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    uid = args.uid

    # Step 1 — Initialize Firebase
    print('Initializing Firebase...')
    db = init_firebase(args.key)

    # Step 2 — Load activity categories
    print('Loading activity categories from Firestore...')
    categories_by_name, parent_by_id = load_categories(db, uid)
    print(f'  {len(categories_by_name)} categories loaded.')

    # Step 3 — Load existing persons
    print('Loading existing persons from Firestore...')
    persons_by_fullname = load_persons(db, uid)
    print(f'  {len(persons_by_fullname)} persons loaded.')

    # Step 4 — Load existing meetings (idempotency)
    print('Loading existing meetings from Firestore...')
    existing_meetings = load_existing_meetings(db, uid)
    print(f'  {len(existing_meetings)} meetings already exist.')

    # Step 5 — Load Excel
    print(f'Loading workbook: {args.xlsx}')
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    person_columns = headers[4:]  # col index 4 onwards
    print(f'  Person columns: {person_columns}')

    # Step 6 — Pre-flight: verify all activity names exist
    preflight_check(ws, headers, categories_by_name)

    # Step 7 — Process rows
    persons_col_ref = db.collection('users').document(uid).collection('persons')
    meetings_col_ref = db.collection('users').document(uid).collection('meetings')

    new_persons: list[tuple] = []        # (doc_id, data_dict)
    new_meetings: list[tuple] = []       # (doc_id, data_dict)
    skipped = 0
    persons_reused = 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    total_rows = len(rows)

    for row in rows:
        # Parse date
        raw_date = row[0]
        if raw_date is None:
            continue
        if isinstance(raw_date, datetime.datetime):
            meeting_date = raw_date.date()
            meeting_datetime = datetime.datetime(
                raw_date.year, raw_date.month, raw_date.day,
                tzinfo=datetime.timezone.utc,
            )
        else:
            # Fallback: already a date object
            meeting_date = raw_date
            meeting_datetime = datetime.datetime(
                raw_date.year, raw_date.month, raw_date.day,
                tzinfo=datetime.timezone.utc,
            )

        # Parse weight
        raw_weight = row[1]
        weight = int(float(raw_weight)) if raw_weight is not None else 1
        weight = WEIGHT_MAP.get(weight, weight)

        # Parse meeting name
        name = str(row[3]).strip() if row[3] is not None else ''
        if not name:
            continue

        # Idempotency check
        if (meeting_date, name) in existing_meetings:
            skipped += 1
            continue

        # Parse activities → category_ids (leaf + ancestors, deduplicated)
        raw_activities = row[2]
        activity_names = [n.strip() for n in str(raw_activities).split(';') if n.strip()] if raw_activities else []
        category_ids = resolve_category_ids(activity_names, categories_by_name, parent_by_id)

        # Parse participants
        participant_ids: list[str] = []
        for col_idx, person_name in enumerate(person_columns):
            cell_value = row[4 + col_idx]
            if cell_value is not None and str(cell_value).strip().lower() == 'x':
                if not person_name:
                    continue
                full_name = str(person_name).strip()
                before_count = len(persons_by_fullname)
                pid = get_or_create_person(full_name, uid, persons_by_fullname, new_persons)
                if len(persons_by_fullname) == before_count:
                    persons_reused += 1
                participant_ids.append(pid)

        now = datetime.datetime.utcnow()
        meeting_doc = {
            'userId': uid,
            'name': name,
            'date': meeting_datetime,
            'weight': weight,
            'participantIds': participant_ids,
            'categoryIds': category_ids,
            'createdAt': now,
            'updatedAt': now,
        }
        new_meetings.append((db_ref_id(), meeting_doc))

    print(f'\nPre-write summary:')
    print(f'  Meetings to import:  {len(new_meetings)}')
    print(f'  Meetings to skip:    {skipped}')
    print(f'  Persons to create:   {len(new_persons)}')
    print(f'  Persons reused:      {persons_reused}')

    if not new_meetings and not new_persons:
        print('\nNothing to write. Done.')
        return

    # Step 8 — Batch write persons first (meetings reference their IDs)
    if new_persons:
        print(f'\nWriting {len(new_persons)} new person(s)...')
        commit_in_batches(db, [(persons_col_ref, doc_id, data) for doc_id, data in new_persons])
        print('  Persons written.')

    # Step 9 — Batch write meetings with progress
    if new_meetings:
        print(f'Writing {len(new_meetings)} new meeting(s)...')
        total = len(new_meetings)
        imported = 0
        for start in range(0, total, BATCH_SIZE):
            chunk = new_meetings[start:start + BATCH_SIZE]
            batch = db.batch()
            for doc_id, data in chunk:
                batch.set(meetings_col_ref.document(doc_id), data)
            batch.commit()
            imported += len(chunk)
            print(f'  Imported {imported}/{total} meetings...')

    # Step 10 — Summary
    print('\nMigration complete.')
    print(f'  Meetings imported: {len(new_meetings)}')
    print(f'  Meetings skipped (already exist): {skipped}')
    print(f'  Persons created: {len(new_persons)}')
    print(f'  Persons reused: {persons_reused}')


if __name__ == '__main__':
    main()
